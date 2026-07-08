#!/usr/bin/env python3
"""
Hilfsskript: wandelt data/Telegram_Wochenplan_mfx555.xlsx neu in data/wochenplan.json um.

Immer ausführen, nachdem die Exceltabelle (Spalten: Tag, Uhrzeit, Slot, Text)
bearbeitet wurde, und das Ergebnis committen/pushen - der Bot liest nur die JSON-Datei.

Benötigt: pip install openpyxl
"""
import datetime
import json
import os

import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_FILE = os.path.join(BASE_DIR, "data", "Telegram_Wochenplan_mfx555.xlsx")
JSON_FILE = os.path.join(BASE_DIR, "data", "wochenplan.json")


def format_uhrzeit(value):
    """Wandelt Excel-Zeitwerte (egal ob Text oder Zeit-Zellformat) zuverlässig in 'HH:MM' um."""
    if isinstance(value, (datetime.time, datetime.datetime)):
        return value.strftime("%H:%M")
    s = str(value).strip()
    # Falls jemand "17:30:00" als Text eingegeben hat: Sekunden abschneiden.
    parts = s.split(":")
    if len(parts) >= 2:
        return f"{int(parts[0]):02d}:{int(parts[1]):02d}"
    return s


def main():
    wb = openpyxl.load_workbook(XLSX_FILE, data_only=True)
    ws = wb["Wochenplan"]

    data = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        tag, uhrzeit, slot, text = [c.value for c in row]
        if tag is None or uhrzeit is None or text is None:
            continue
        data.append({
            "tag": str(tag).strip().upper(),
            "uhrzeit": format_uhrzeit(uhrzeit),
            "slot": str(slot).strip() if slot else "",
            "text": str(text),
        })

    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"{len(data)} Zeilen aus {XLSX_FILE} nach {JSON_FILE} exportiert.")


if __name__ == "__main__":
    main()
